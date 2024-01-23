from openpyxl import load_workbook
from openpyxl.styles import Font
import datetime
from copy import copy
import calendar
import os

template_workbook = load_workbook(os.path.expanduser('~/Documents/Excel_templates/Templates.xlsx'))
template_worksheets = template_workbook.active

output_workbook = load_workbook(os.path.expanduser('~/Documents/Excel_templates/Output_Workbook.xlsx'))
output_worksheets = template_workbook.active

curr_date = datetime.date.today()
first = curr_date.replace(day=1)
last_month = first - datetime.timedelta(days=1)
last_month = datetime.date(last_month.year, last_month.month, 1)

sheet_name = f'Salary_Report_{last_month.strftime("%m-%Y")}'

here = False
for names in output_workbook.sheetnames:
    if names == sheet_name:
        here = True
        break
if not here:
    output_workbook.create_sheet(sheet_name)

merge_coordinates = {1: [('A', 'L')],
                     2: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     3: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     4: [('A', 'L')],
                     5: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     6: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     7: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     8: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     9: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     10: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     11: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     12: [('A', 'L')],
                     13: [('A', 'L')],
                     14: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     15: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     16: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     17: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     18: [('A', 'E'), ('F', 'L')]
                     }


def Copy_Output_Template(offset: int) -> None:
    for _ in range(2):
        curr_row = 1
        for _ in template_workbook['Empty_template'].rows:
            for left, right in merge_coordinates[curr_row]:
                # print(curr_row, left, right)
                begin_coordinates, end_coordinates = f'{left}{curr_row + offset}', f'{right}{curr_row + offset}'
                another_used_begin = f'{left}{curr_row}'
                output_workbook[sheet_name].merge_cells(f'{begin_coordinates}:{end_coordinates}')
                template_cell = template_workbook['Empty_template'][f'{another_used_begin}']
                curr_value = template_cell.value
                new_cell = output_workbook[sheet_name].cell(row=curr_row + offset, column=ord(left) - 64,
                                                            value=curr_value)
                # print(f'Value: {new_cell.value}')
                new_cell.font = copy(template_cell.font)
                new_cell.border = copy(template_cell.border)
                new_cell.alignment = copy(template_cell.alignment)
                new_cell.fill = copy(template_cell.fill)
            curr_row += 1
        output_workbook.save(os.path.expanduser('~/Documents/Excel_templates/Output_Workbook.xlsx'))


def Process_Input_Template() -> None:
    curr_row, offset = 2, 0
    input_worksheet = template_workbook['Input_template']
    output_worksheet = output_workbook[sheet_name]
    while curr_row < 16 and input_worksheet[f'A{curr_row}'].value is not None:

        resigning = False

        print(f'Processing row {curr_row}')

        # Create the template
        Copy_Output_Template(offset)

        # Reading user inputs
        name = input_worksheet[f'A{curr_row}'].value
        basic_salary = input_worksheet[f'C{curr_row}'].value
        join_date = input_worksheet[f'E{curr_row}'].value
        sick_leave_taken = input_worksheet[f'G{curr_row}'].value
        no_paid_leave_taken = input_worksheet[f'I{curr_row}'].value
        resigning_date = input_worksheet[f'K{curr_row}'].value
        if resigning_date is not None:
            resigning = True
        annual_leave = input_worksheet[f'M{curr_row}'].value
        cheque_number = input_worksheet[f'O{curr_row}'].value
        paid_date = input_worksheet[f'Q{curr_row}'].value

        if basic_salary is None or sick_leave_taken is None or no_paid_leave_taken is None or (
                resigning and annual_leave is None):
            return

        # Calculate probation due date
        day, month, year = join_date.day, join_date.month, join_date.year
        month += 3
        if month > 12:
            month -= 12
            year += 1

        last_day = calendar.monthrange(year, month)[1]
        if day > last_day:
            day -= last_day
            month += 1
            if month > 12:
                month -= 12
                year += 1
        probation_due = datetime.date(year, month, day)
        remind = -1
        messages = [f'On going probation, all {sick_leave_taken + no_paid_leave_taken} leaves deducted',
                    f'Caution: Probation ended at {probation_due.strftime("%d-%m-%Y")}. Revise total deduction']

        # Calculate leave taken salary deduction
        days_at_work = calendar.monthrange(curr_date.year, last_month.month)[1] if (
                    join_date.month != curr_date.month - 1 or
                    (join_date.month == last_month.month and join_date.year != curr_date.year)) else \
        calendar.monthrange(curr_date.year, last_month.month)[1] - join_date.day + 1
        number_of_days = calendar.monthrange(curr_date.year, last_month.month)[1] if last_month.month != 2 else 28
        accurate_day_salary = basic_salary / number_of_days
        total_deduction, paid_leave_salary, alternate = 0, 0, 0

        temp = datetime.date(paid_date.year, paid_date.month, paid_date.day)

        if last_month >= probation_due or last_month < probation_due < temp:
            if sick_leave_taken > 0:
                total_deduction += accurate_day_salary * (sick_leave_taken - 1) if sick_leave_taken - 1 > 0 else 0
                paid_leave_salary += accurate_day_salary if sick_leave_taken >= 1 else accurate_day_salary * 0.5
            total_deduction += accurate_day_salary * no_paid_leave_taken
            if last_month < probation_due < temp:
                remind = 1

        elif temp <= probation_due:
            total_leave = sick_leave_taken + no_paid_leave_taken
            total_deduction += (accurate_day_salary * total_leave)
            remind = 0

        # Resigning process (annual leave taken)
        if resigning:
            if resigning_date.year == join_date.year:
                start_date = join_date
            else:
                start_date = datetime.datetime(resigning_date.year, 1, 1)
            date_delta = resigning_date - start_date
            annual_limit = round((date_delta.days + 1) / 365 * 10, 1)
            if annual_leave <= annual_limit:
                alternate = round((annual_limit - annual_leave) * accurate_day_salary, 2)
            else:
                extra_days = annual_leave - annual_limit
                total_deduction += extra_days * accurate_day_salary

        temp_net_payment = accurate_day_salary * days_at_work + alternate - total_deduction

        # MPF
        Employer_MPF_Contribution = round(temp_net_payment * 0.05, 2) if basic_salary < 30000 else 1500
        Employee_MPF_Contribution = 0
        date_diff = paid_date - join_date
        if date_diff.days + 1 >= 60:
            Employee_MPF_Contribution = round(temp_net_payment * 0.05,
                                              2) if basic_salary < 30000 else 1500

        total_deduction += Employee_MPF_Contribution

        total_deduction = round(total_deduction, 2)

        net_paid = round(accurate_day_salary * days_at_work + alternate - total_deduction, 2)

        # Putting the data to the corresponding places
        output_worksheet[f'D{2 + offset}'] = name
        output_worksheet[f'E{6 + offset}'] = basic_salary
        output_worksheet[f'J{2 + offset}'] = resigning_date.strftime('%d-%m-%Y') if resigning else None
        output_worksheet[f'K{6 + offset}'] = annual_leave
        output_worksheet[f'K{7 + offset}'] = no_paid_leave_taken
        if remind == -1 or remind == 1:
            output_worksheet[f'K{8 + offset}'] = sick_leave_taken - 1 if sick_leave_taken - 1 > 0 else 0
        else:
            output_worksheet[f'K{8 + offset}'] = sick_leave_taken
        output_worksheet[f'K{9 + offset}'] = Employee_MPF_Contribution
        output_worksheet[f'K{10 + offset}'] = total_deduction
        output_worksheet[f'K{11 + offset}'] = net_paid
        if remind != -1:
            output_worksheet[f'A{12 + offset}'].font = Font(color='b7312c', name='Calibri', bold=True)
            output_worksheet[f'A{12 + offset}'] = messages[remind]
        if remind == -1 or remind == 1:
            output_worksheet[f'E{7 + offset}'] = 1 if sick_leave_taken >= 1 else sick_leave_taken
        else:
            output_worksheet[f'E{7 + offset}'] = 0
        output_worksheet[f'E{8 + offset}'] = alternate
        output_worksheet[f'E{10 + offset}'] = round(alternate, 2)
        output_worksheet[f'K{15 + offset}'] = Employer_MPF_Contribution
        output_worksheet[f'K{16 + offset}'] = Employee_MPF_Contribution
        output_worksheet[f'D{17 + offset}'] = cheque_number
        output_worksheet[f'J{17 + offset}'] = paid_date.strftime('%d-%m-%Y')
        output_worksheet[f'F{18 + offset}'] = curr_date
        curr_row += 1
        offset += 20
    print('Finished processing')
    output_workbook.save(os.path.expanduser('~/Documents/Excel_templates/Output_Workbook.xlsx'))


Process_Input_Template()
