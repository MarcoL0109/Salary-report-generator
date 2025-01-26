from openpyxl import load_workbook
from openpyxl.styles import Font
import os
from copy import copy
import datetime



class Excel_Processor():
    
    def __init__(self, template_workbook, output_workbook):
        self.template_workbook = template_workbook
        self.output_workbook = output_workbook
        self.template_workbook.active
        self.output_workbook.active
        self.merge_coordinates = {1: [('A', 'L')],
                     2: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     3: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     4: [('A', 'L')],
                     5: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     6: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     7: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     8: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     9: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     10: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     11: [('A', 'D'), ('E', 'F'), ('G', 'J'), ('K', 'L')],
                     12: [('A', 'L')],
                     13: [('A', 'L')],
                     14: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     15: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     16: [('A', 'H'), ('I', 'J'), ('K', 'L')],
                     17: [('A', 'C'), ('D', 'F'), ('G', 'I'), ('J', 'L')],
                     18: [('A', 'E'), ('F', 'L')]
                     }
        curr_date = datetime.date.today()
        first = curr_date.replace(day=1)
        last_month = first - datetime.timedelta(days=1)
        self.last_month = datetime.date(last_month.year, last_month.month, 1)
        self.sheet_name = f'Salary_Report_{last_month.strftime("%m-%Y")}'
        self.input_worksheet = self.template_workbook['Input_template']

        here = False
        for names in self.output_workbook.sheetnames:
            if names == self.sheet_name:
                here = True
                break
        if not here:
            self.output_workbook.create_sheet(self.sheet_name)
        
        self.output_worksheet = self.output_workbook[self.sheet_name]


    def Copy_Output_Template(self, offset: int) -> None:
        for _ in range(2):
            curr_row = 1
            for _ in self.template_workbook['Empty_template'].rows:
                for left, right in self.merge_coordinates[curr_row]:
                    begin_coordinates, end_coordinates = f'{left}{curr_row + offset}', f'{right}{curr_row + offset}'
                    another_used_begin = f'{left}{curr_row}'
                    self.output_workbook[self.sheet_name].merge_cells(f'{begin_coordinates}:{end_coordinates}')
                    template_cell = self.template_workbook['Empty_template'][f'{another_used_begin}']
                    curr_value = template_cell.value
                    new_cell = self.output_workbook[self.sheet_name].cell(row=curr_row + offset, column=ord(left) - 64,
                                                                value=curr_value)
                    new_cell.font = copy(template_cell.font)
                    new_cell.border = copy(template_cell.border)
                    new_cell.alignment = copy(template_cell.alignment)
                    new_cell.fill = copy(template_cell.fill)
                curr_row += 1
            self.output_workbook.save(os.path.expanduser('Salary_Excel_Generator/Excel_templates/Output_Workbook.xlsx'))
    

    def place_values_to_output(self, name, basic_salary, probation_due, resigning, resigning_date, annual_leave, no_paid_leave_taken, sick_leave_taken, Employee_MPF_Contribution, Employer_MPF_Contribution, total_deduction, net_paid, alternate, cheque_number, paid_date, curr_date, remind, offset) -> None:
        messages = [f'On going probation, all {sick_leave_taken + no_paid_leave_taken} leaves deducted',
                    f'Caution: Probation ended at {probation_due.strftime("%d-%m-%Y")}. Revise total deduction']
        self.output_worksheet[f'D{2 + offset}'] = name
        self.output_worksheet[f'E{6 + offset}'] = basic_salary
        self.output_worksheet[f'J{2 + offset}'] = resigning_date.strftime('%d-%m-%Y') if resigning else None
        self.output_worksheet[f'K{6 + offset}'] = annual_leave
        self.output_worksheet[f'K{7 + offset}'] = no_paid_leave_taken
        if remind == -1 or remind == 1:
            self.output_worksheet[f'K{8 + offset}'] = sick_leave_taken - 1 if sick_leave_taken - 1 > 0 else 0
        else:
            self.output_worksheet[f'K{8 + offset}'] = sick_leave_taken
        self.output_worksheet[f'K{9 + offset}'] = Employee_MPF_Contribution
        self.output_worksheet[f'K{10 + offset}'] = total_deduction
        self.output_worksheet[f'K{11 + offset}'] = net_paid
        if remind != -1:
            self.output_worksheet[f'A{12 + offset}'].font = Font(color='b7312c', name='Calibri', bold=True)
            self.output_worksheet[f'A{12 + offset}'] = messages[remind]
        else:
            self.output_worksheet[f'A{12 + offset}'] = None
        if remind == -1 or remind == 1:
            self.output_worksheet[f'E{7 + offset}'] = 1 if sick_leave_taken >= 1 else sick_leave_taken
        else:
            self.output_worksheet[f'E{7 + offset}'] = 0
        self.output_worksheet[f'E{8 + offset}'] = alternate
        self.output_worksheet[f'E{10 + offset}'] = round(alternate, 2)
        self.output_worksheet[f'K{15 + offset}'] = Employer_MPF_Contribution
        self.output_worksheet[f'K{16 + offset}'] = Employee_MPF_Contribution
        self.output_worksheet[f'D{17 + offset}'] = cheque_number
        self.output_worksheet[f'J{17 + offset}'] = paid_date.strftime('%d-%m-%Y')
        self.output_worksheet[f'F{18 + offset}'] = curr_date
    

    def get_info(self, curr_row):
        name = self.input_worksheet[f'A{curr_row}'].value
        basic_salary = self.input_worksheet[f'C{curr_row}'].value
        join_date = self.input_worksheet[f'E{curr_row}'].value
        sick_leave_taken = self.input_worksheet[f'G{curr_row}'].value
        no_paid_leave_taken = self.input_worksheet[f'I{curr_row}'].value
        resigning_date = self.input_worksheet[f'K{curr_row}'].value
        resigning = True if resigning_date is not None else False
        annual_leave = self.input_worksheet[f'M{curr_row}'].value
        cheque_number = self.input_worksheet[f'O{curr_row}'].value
        paid_date = self.input_worksheet[f'Q{curr_row}'].value

        return name, basic_salary, join_date, sick_leave_taken, no_paid_leave_taken, resigning_date, resigning, annual_leave, cheque_number, paid_date
    
    
