from calculator.calculator import Calculator
from excel_process.excel_process import Excel_Processor
from openpyxl import load_workbook
import os




def main():

    template_path = 'Salary_Excel_Generator/Excel_templates/Templates.xlsx'
    output_path = 'Salary_Excel_Generator/Excel_templates/Output_Workbook.xlsx'
    template_workbook = load_workbook(os.path.expanduser(template_path))
    output_workbook = load_workbook(os.path.expanduser(output_path))

    excel_processor = Excel_Processor(template_workbook, output_workbook)
    calculator = Calculator()
    input_worksheet = template_workbook['Input_template']

    curr_row, offsets = 2, 0
    while input_worksheet[f'A{curr_row}'].value is not None:
        print(f'Processing row {curr_row} with name {input_worksheet[f"A{curr_row}"].value}')
        excel_processor.Copy_Output_Template(offsets)
        name, basic_salary, join_date, sick_leave_taken, no_paid_leave_taken, resigning_date, resigning, annual_leave, cheque_number, paid_date = excel_processor.get_info(curr_row)
        calculator.update_values(basic_salary, join_date, sick_leave_taken, no_paid_leave_taken, resigning_date, resigning, annual_leave, paid_date)

        probation_due = calculator.find_probation()
        leave_taken_deduction, remind = calculator.find_leave_taken_deduction(probation_due)
        resign_deduction, alternate = calculator.find_resign_deduction()
        total_deduction = leave_taken_deduction + resign_deduction
        Employee_MPF_Contribution, Employer_MPF_Contribution = calculator.MPF_calculation(alternate, total_deduction)
        total_deduction += Employee_MPF_Contribution
        final_net_paid = calculator.find_final_net_payment(alternate, total_deduction)

        excel_processor.place_values_to_output(name, basic_salary, probation_due, resigning, resigning_date, annual_leave, no_paid_leave_taken, sick_leave_taken, Employee_MPF_Contribution, Employer_MPF_Contribution, total_deduction, final_net_paid, alternate, cheque_number, paid_date, calculator.curr_date, remind, offsets)
        curr_row += 1
        offsets += 20
    
    output_workbook.save(os.path.expanduser(output_path))


if __name__ == '__main__':
    main()